import io
import re
import base64
from collections import defaultdict
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Streamlit chrome
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="uKids — Kids Scheduler (Matches Your CSVs)", layout="wide")
st.title("uKids — Kids Scheduler (Matches Your CSVs)")

st.markdown(
    """
    <style>
      .stApp { background: #000; color: #fff; }
      .stButton>button, .stDownloadButton>button { background:#444; color:#fff; }
      .stDataFrame { background:#111; }
      .stAlert { color:#fff; }
      a, .stDownloadButton>button:hover, .stButton>button:hover { filter: brightness(1.1); }
      .hint { opacity: .8; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Optional logo (ignore if missing)
for logo_name in ["image(1).png", "image.png", "logo.png"]:
    try:
        with open(logo_name, "rb") as f:
            encoded = base64.b64encode(f.read()).decode()
        st.markdown(
            f"<div style='text-align:center'><img src='data:image/png;base64,{encoded}' width='420'></div>",
            unsafe_allow_html=True,
        )
        break
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# Helpers & parsing
# ──────────────────────────────────────────────────────────────────────────────
MONTH_ALIASES = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
YES_SET = {"yes", "y", "true", "available", "1", "x", "✓", "ok", "okay"}

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()

def read_csv_robust(uploaded_file, label_for_error):
    raw = uploaded_file.getvalue()
    encodings = ["utf-8", "utf-8-sig", "cp1252", "iso-8859-1"]
    seps = [None, ",", ";", "\t", "|"]
    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, engine="python", sep=sep)
                if df.shape[1] == 0:
                    raise ValueError("Parsed 0 columns.")
                return df
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                continue
    st.error(
        f"Could not read {label_for_error} CSV. Last error: {last_err}. "
        "Try re-exporting as CSV (UTF-8) or remove unusual characters in headers."
    )
    st.stop()

CAP_PATTERNS = [
    re.compile(r"^(?P<base>.*?)[\s\-]*\(\s*x?\s*(?P<n>\d+)\s*\)\s*$", re.IGNORECASE),
    re.compile(r"^(?P<base>.*?)[\s\-]*\[\s*x?\s*(?P<n>\d+)\s*\]\s*$", re.IGNORECASE),
    re.compile(r"^(?P<base>.*?)[\s\-]*x\s*(?P<n>\d+)\s*$", re.IGNORECASE),
]

def parse_role_meta(header: str, default_cap: int):
    """Return (base_label, session, capacity:int) from a header like 'Kids Production M (x4)'. Session is 'Morning'/'Evening'/'Both'."""
    s = str(header).replace("\u00A0", " ").replace("×", "x").strip()
    base = s
    cap = None
    for pat in CAP_PATTERNS:
        m = pat.match(s)
        if m:
            base = str(m.group("base")).strip()
            cap = int(m.group("n"))
            break
    if cap is None:
        cap = default_cap

    nb = normalize(base)
    session = "Both"
    label = base
    if nb.endswith(" m") or nb.endswith(" morning"):
        session = "Morning"
        label = base.rsplit(" ", 1)[0] if " " in base else base
    elif nb.endswith(" e") or nb.endswith(" evening"):
        session = "Evening"
        label = base.rsplit(" ", 1)[0] if " " in base else base
    return label.strip(), session, cap

def truthy_cell(x) -> bool:
    s = str(x).strip().lower()
    if s == "" or s == "nan":
        return False
    return (s in YES_SET)

# ──────────────────────────────────────────────────────────────────────────────
# Date/session parsing from Kids Availability CSV
# ──────────────────────────────────────────────────────────────────────────────
def detect_date_session_headers(df: pd.DataFrame):
    """
    Parse columns like '5 Oct Morning' / '05-Oct Evening' / '19 Oct M' / '26 Oct E'.
    Returns:
      - date_map: {original_col: (pd.Timestamp(date), session_str)}
      - service_slots: sorted unique list of (date, session_str)
      - sheet_name: 'Month YYYY'
    """
    pat = re.compile(r"\b(\d{1,2})\s*[-/ ]\s*([A-Za-z]{3,})\b(?:\s+([A-Za-z]{1,20}))?", re.IGNORECASE)
    avail_cols = [c for c in df.columns if isinstance(c, str) and pat.search(c)]
    if not avail_cols:
        raise ValueError("No availability columns found. Use headers like '05-Oct Morning' or '5 Oct Evening'.")

    def norm_session(tok: str) -> str:
        t = (tok or "").strip().lower()
        if t in ("am", "morning", "m", "1st", "first", "service1", "service-1", "service"):
            return "Morning"
        if t in ("pm", "evening", "e", "2nd", "second", "service2", "service-2"):
            return "Evening"
        return "Service" if t == "" else tok.title()

    info, months_found = [], set()
    for c in avail_cols:
        m = pat.search(str(c))
        if not m:
            continue
        day = int(m.group(1))
        mon_txt = m.group(2).lower()[:3]
        month = MONTH_ALIASES.get(mon_txt)
        session = norm_session(m.group(3))
        months_found.add(month)
        info.append((c, month, day, session))

    if not months_found or None in months_found:
        raise ValueError("Could not parse month from availability headers.")
    if len(months_found) > 1:
        raise ValueError(f"Multiple months detected in availability headers: {sorted(months_found)}. Upload one month at a time.")
    month = months_found.pop()

    if "Timestamp" in df.columns:
        years = pd.to_datetime(df["Timestamp"], errors="coerce").dt.year.dropna().astype(int)
        year = int(years.mode().iloc[0]) if not years.empty else date.today().year
    else:
        year = date.today().year

    date_map, service_slots, seen = {}, [], set()
    for c, mnum, d, session in info:
        dt = pd.Timestamp(datetime(year, mnum, d)).normalize()
        slot = (dt, session)
        date_map[c] = slot
        if slot not in seen:
            seen.add(slot)
            service_slots.append(slot)

    service_slots.sort(key=lambda t: (t[0], t[1]))
    sheet_name = f"{pd.Timestamp(year=year, month=month, day=1):%B %Y}"
    return date_map, service_slots, sheet_name

# ──────────────────────────────────────────────────────────────────────────────
# UI — file inputs & settings
# ──────────────────────────────────────────────────────────────────────────────
st.subheader("1) Upload files")
c1, c2 = st.columns(2)
with c1:
    positions_file = st.file_uploader("Kids Serving Positions (CSV)", type=["csv"], key="positions_csv")
with c2:
    responses_file = st.file_uploader("Kids responses (CSV)", type=["csv"], key="responses_csv")

st.caption("• Positions CSV *like yours*: has 'Kids name', 'Parent Name', then columns per position, e.g. 'Kids Production M', 'Kids Production E', ... Optional capacity in headers: (xN).")
st.caption("• Responses CSV: has 'Kids name' + availability columns like '5 October Morning', '5 October Evening', '12 Oct Morning', '12 Oct Evening', ...")

st.subheader("2) Rules")
default_capacity = st.number_input("Default capacity for positions WITHOUT (xN)", min_value=1, max_value=50, value=3, step=1)
max_serves = st.number_input("Max serves per kid (this month)", min_value=1, max_value=20, value=4, step=1)

# Optional Evening Allowed override
st.caption("Tip: Add an 'Evening Allowed' column in Responses CSV with Yes/No to block evenings for specific kids (defaults to Yes if missing).")

# ──────────────────────────────────────────────────────────────────────────────
# Scheduling logic (positions + sessions) — **matches your CSVs**
# ──────────────────────────────────────────────────────────────────────────────
def excel_autofit(ws):
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1
    ):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 80)

def detect_name_cols(cols):
    cols_map = {normalize(c): c for c in cols}
    kid_name = cols_map.get("kids name") or cols_map.get("name") or list(cols)[0]
    parent_name = cols_map.get("parent name")
    return kid_name, parent_name

def parse_positions_and_eligibility(positions_df: pd.DataFrame, default_cap: int):
    """
    Your Positions CSV has kid rows and many columns for positions.
    We:
      - Detect kid name column
      - Identify position columns (end with ' M'/' E' or 'Morning'/'Evening')
      - Read capacity from header (xN) else default_cap
      - Build per-kid eligibility for each session-specific position
    Returns:
      positions: list of (display_label, session, capacity)
      eligibility: {kid_key: set(display_label_for_session)}  # e.g., 'Kids Production (Morning)'
      name_map: {kid_key: display_name}
      parent_map: {kid_key: parent_name or ""}
    """
    kid_name_col, parent_name_col = detect_name_cols(positions_df.columns)

    # Which columns are positions?
    pos_cols = []
    for c in positions_df.columns:
        if c == kid_name_col or c == parent_name_col:
            continue
        s = str(c).strip()
        if re.search(r"(?:\s|^)(M|E|Morning|Evening)\s*(?:\(|$)", s, re.IGNORECASE) or s.lower().endswith((" m", " e", " morning", " evening")):
            pos_cols.append(c)
        else:
            # ignore other columns silently
            pass

    # Build positions list (unique by base label + session)
    positions = []
    seen_pos = set()
    for c in pos_cols:
        label, sess, cap = parse_role_meta(c, default_cap)
        # Create display label with explicit session for clarity
        display = f"{label} ({'Morning' if sess=='Morning' else 'Evening' if sess=='Evening' else 'Both'})"
        # If 'Both', split into two session-specific displays for roster rows
        if sess == "Both":
            for s2 in ("Morning", "Evening"):
                key = (f"{label} ({s2})", s2, cap)
                if key not in seen_pos:
                    positions.append(key)
                    seen_pos.add(key)
        else:
            key = (f"{label} ({sess})", sess, cap)
            if key not in seen_pos:
                positions.append(key)
                seen_pos.add(key)

    # Per-kid eligibility
    def norm_session_token(h):
        nb = normalize(h)
        if nb.endswith(" m") or nb.endswith(" morning"):
            return "Morning"
        if nb.endswith(" e") or nb.endswith(" evening"):
            return "Evening"
        return "Both"

    name_map, parent_map, eligibility = {}, {}, defaultdict(set)
    for _, row in positions_df.iterrows():
        kid_disp = str(row.get(kid_name_col, "")).strip()
        if not kid_disp:
            continue
        key = normalize(kid_disp)
        name_map[key] = kid_disp
        parent_map[key] = (str(row.get(parent_name_col, "")).strip() if parent_name_col else "")

        # For each position column, if cell truthy → eligible for that session
        for c in pos_cols:
            if not truthy_cell(row.get(c, "")):
                continue
            label, sess, _cap = parse_role_meta(c, default_cap)
            if sess == "Both":
                eligibility[key].add(f"{label} (Morning)")
                eligibility[key].add(f"{label} (Evening)")
            else:
                eligibility[key].add(f"{label} ({sess})")

    # Sort positions by name then session
    positions.sort(key=lambda t: (normalize(t[0]), t[1]))
    return positions, eligibility, name_map, parent_map

def parse_availability(responses_df: pd.DataFrame):
    # Name column(s)
    kid_name_col, parent_name_col = detect_name_cols(responses_df.columns)

    # Optional Evening Allowed in responses
    cols_map = {normalize(c): c for c in responses_df.columns}
    evening_allowed_col = cols_map.get("evening allowed") or cols_map.get("evening_allowed") or cols_map.get("allowed evening")

    # Service slots from headers
    date_map, service_slots, sheet_name = detect_date_session_headers(responses_df)

    # Build availability & name maps from responses
    avail = defaultdict(dict)
    name_map = {}
    evening_allowed = {}

    for _, row in responses_df.iterrows():
        kid_disp = str(row.get(kid_name_col, "")).strip()
        if not kid_disp:
            continue
        key = normalize(kid_disp)
        name_map[key] = kid_disp
        if evening_allowed_col:
            evening_allowed[key] = str(row.get(evening_allowed_col, "yes")).strip().lower() in YES_SET
        else:
            evening_allowed[key] = True

        for col, slot in date_map.items():
            ans = str(row.get(col, "")).strip().lower()
            avail[key][slot] = (ans in YES_SET)

    return avail, evening_allowed, name_map, service_slots, sheet_name

def assign_kids(positions, eligibility, availability, evening_allowed, service_slots, max_serves=4):
    """
    positions: list of (display_label, session, capacity) — session is 'Morning' or 'Evening'
    eligibility: {kid_key: set(display_label_for_session)}
    availability: {kid_key: {(date, session): bool}}
    evening_allowed: {kid_key: bool}
    service_slots: list of (date, session)

    Rules:
      - Kid must be eligible for that position/session
      - Kid must be available for that (date, session)
      - Max total assignments per kid = max_serves
      - Never assign the same kid twice on the same calendar date (across sessions/positions)
      - If Evening Allowed is False → cannot take Evening slots
    """
    # Build quick kid list (union from eligibility & availability)
    kid_keys = sorted(set(list(eligibility.keys()) + list(availability.keys())))
    # (position_display, (date, session)) → [kid display names]
    schedule = {(disp, slot): [] for (disp, sess, _c) in positions for slot in service_slots if slot[1] == sess}
    unscheduled = {slot: [] for slot in service_slots}

    assigned_total = defaultdict(int)   # per month
    assigned_on_date = set()            # (kid_key, date)

    # Fairness: rotate by (assigned_total, name)
    def kid_sort_key(k): return (assigned_total[k], k)

    for slot in service_slots:
        d, sess = slot
        assigned_this_slot = set()

        # Positions for this session
        session_positions = [(disp, sess2, cap) for (disp, sess2, cap) in positions if sess2 == sess]

        for (disp, sess2, cap) in session_positions:
            # candidates: eligible + available + evening ok + not already serving date + under cap
            candidates = []
            for key in kid_keys:
                if key in assigned_this_slot:            # slot-unique
                    continue
                if assigned_total[key] >= max_serves:    # monthly cap
                    continue
                if (key, d) in assigned_on_date:         # not twice same day
                    continue
                if disp not in eligibility.get(key, set()):
                    continue
                if not availability.get(key, {}).get(slot, False):
                    continue
                if sess == "Evening" and not evening_allowed.get(key, True):
                    continue
                candidates.append(key)

            # Fairness order
            candidates.sort(key=kid_sort_key)

            while candidates and len(schedule[(disp, slot)]) < cap:
                chosen = candidates.pop(0)
                schedule[(disp, slot)].append(chosen)    # temporarily store key; convert to names later
                assigned_this_slot.add(chosen)
                assigned_total[chosen] += 1
                assigned_on_date.add((chosen, d))

        # Anyone who said YES for this slot but not assigned → unscheduled
        for key in [k for k in kid_keys if availability.get(k, {}).get(slot, False)]:
            if key not in assigned_this_slot:
                unscheduled[slot].append(key)

    return schedule, unscheduled

def schedule_to_df(schedule, service_slots, name_map):
    cols = [f"{d.strftime('%Y-%m-%d')} ({s})" for (d, s) in service_slots]
    rows = sorted({disp for (disp, _slot) in schedule.keys()})
    df = pd.DataFrame(index=rows, columns=cols)
    for (disp, slot), kids in schedule.items():
        d, s = slot
        names = [name_map.get(k, k) for k in kids]
        df.loc[disp, f"{d.strftime('%Y-%m-%d')} ({s})"] = ", ".join(names)
    return df.fillna("")

def unscheduled_to_df(unscheduled, service_slots, name_map):
    per_slot_names = {}
    for (d, s) in service_slots:
        keys = sorted(unscheduled.get((d, s), []))
        per_slot_names[(d, s)] = [name_map.get(k, k) for k in keys]

    max_len = max((len(v) for v in per_slot_names.values()), default=0)
    data = {}
    for (d, s) in service_slots:
        key = f"{d.strftime('%Y-%m-%d')} ({s})"
        data[key] = per_slot_names[(d, s)] + [""] * (max_len - len(per_slot_names[(d, s)]))
    return pd.DataFrame(data)

# ──────────────────────────────────────────────────────────────────────────────
# Run
# ──────────────────────────────────────────────────────────────────────────────
if st.button("Generate Kids Schedule", type="primary"):
    if not positions_file or not responses_file:
        st.error("Please upload the Positions and Responses CSV files.")
        st.stop()

    positions_df = read_csv_robust(positions_file, "positions")
    responses_df = read_csv_robust(responses_file, "responses")

    # Parse positions + eligibility from your Positions CSV
    positions, eligibility, name_map_positions, parent_map = parse_positions_and_eligibility(positions_df, default_cap=default_capacity)

    # Parse availability from your Responses CSV
    availability, evening_allowed, name_map_responses, service_slots, sheet_name = parse_availability(responses_df)

    # Merge display names (prefer positions casing; fall back to responses)
    name_map = {**name_map_responses, **name_map_positions, **name_map_responses}

    # Assign with rules
    schedule, unscheduled = assign_kids(
        positions, eligibility, availability, evening_allowed, service_slots, max_serves=max_serves
    )

    schedule_df = schedule_to_df(schedule, service_slots, name_map)
    unscheduled_df = unscheduled_to_df(unscheduled, service_slots, name_map)

    # Stats (cells that have at least one kid)
    filled_cells = sum(1 for v in schedule.values() if len(v) > 0)
    total_valid_cells = len(schedule)
    total_kids_scheduled = sum(len(v) for v in schedule.values())

    st.success(f"Schedule generated for **{sheet_name}**")
    st.write(f"Kids scheduled: **{total_kids_scheduled}**  •  Positions with any kids: **{filled_cells} / {total_valid_cells}**")

    st.subheader("Rosters (Position × Service)")
    st.dataframe(schedule_df, use_container_width=True)

    st.subheader("Unscheduled but Available — by Service")
    st.caption("Kids who said Yes for that service but were not placed (capacity reached, daily rule, evening rule, or monthly cap).")
    st.dataframe(unscheduled_df, use_container_width=True)

    # Excel export
    wb = Workbook()
    if wb.active and wb.active.title == "Sheet":
        wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)

    header = ["Position"] + [f"{d.strftime('%Y-%m-%d')} ({s})" for (d, s) in service_slots]
    ws.append(header)
    row_labels = sorted({disp for (disp, _slot) in schedule.keys()})
    for disp in row_labels:
        vals = []
        for (d, s) in service_slots:
            names = [name_map.get(k, k) for k in schedule.get((disp, (d, s)), [])]
            vals.append(", ".join(names))
        ws.append([disp] + vals)
    excel_autofit(ws)

    ws2 = wb.create_sheet("Unscheduled by Service")
    ws2.append([" "] + [f"{d.strftime('%Y-%m-%d')} ({s})" for (d, s) in service_slots])
    for i in range(unscheduled_df.shape[0]):
        ws2.append([i + 1] + [unscheduled_df.iloc[i, j] for j in range(unscheduled_df.shape[1])])
    excel_autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "Download Excel (.xlsx)",
        data=buf,
        file_name=f"uKids_Kids_Schedule_{sheet_name.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload your **Kids Serving Positions** CSV and **Kids responses** CSV, set rules, then click **Generate Kids Schedule**.")

